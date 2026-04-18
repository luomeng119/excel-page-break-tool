#!/usr/bin/env python3
"""
卷内目录分页处理器 - Python 后端（可选）

功能：当顺序号=1时，在该行前插入分页符。

此文件是可选后端，前端 index.html 已包含完整功能（纯浏览器端运行）。
后端适用于以下场景：
  - 需要服务端批量处理大量文件
  - 需要精确控制 Excel 打印样式
  - 需要集成到现有 Flask/FastAPI 服务中

依赖：pip install flask flask-cors openpyxl
启动：python processor.py
"""

import io
import openpyxl
from openpyxl.worksheet.pagebreak import Break
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

app = Flask(__name__)
CORS(app)


def find_page_breaks(data_rows, seq_col_index=0):
    breaks = []
    for i in range(1, len(data_rows)):
        try:
            curr_val = int(str(data_rows[i][seq_col_index]).strip())
            prev_val = int(str(data_rows[i - 1][seq_col_index]).strip())
        except (ValueError, TypeError, AttributeError):
            continue
        if curr_val == 1 and prev_val != 1:
            breaks.append(i)
    return breaks


def analyze_excel(file_stream):
    wb = openpyxl.load_workbook(file_stream, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_rows = 2
    data_rows = all_rows[header_rows:]
    breaks = find_page_breaks(data_rows, seq_col_index=0)

    volumes = []
    start = 0
    for b in breaks:
        volumes.append({
            "volumeNo": len(volumes) + 1,
            "recordCount": b - start,
            "startRow": start + header_rows + 1,
        })
        start = b
    volumes.append({
        "volumeNo": len(volumes) + 1,
        "recordCount": len(data_rows) - start,
        "startRow": start + header_rows + 1,
    })

    return {
        "total_rows": len(all_rows),
        "data_rows": len(data_rows),
        "volumes_count": len(volumes),
        "page_breaks_count": len(breaks),
        "volumes": volumes,
    }


def process_excel_with_pagebreaks(file_stream, filename):
    wb = openpyxl.load_workbook(file_stream)
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    header_rows = 2
    data_rows = all_rows[header_rows:]
    breaks = find_page_breaks(data_rows, seq_col_index=0)

    for b in breaks:
        excel_row = b + header_rows + 1
        ws.row_breaks.append(Break(id=excel_row))

    volumes = []
    start = 0
    for b in breaks:
        volumes.append({"volumeNo": len(volumes) + 1, "recordCount": b - start})
        start = b
    volumes.append({"volumeNo": len(volumes) + 1, "recordCount": len(data_rows) - start})

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)

    stats = {
        "total_rows": len(all_rows),
        "data_rows": len(data_rows),
        "volumes_count": len(volumes),
        "page_breaks_count": len(breaks),
        "volumes": volumes,
    }
    return output, stats


@app.route("/api/health", methods=["GET"])
def api_health():
    return jsonify({"status": "ok", "version": "1.0.0"})


@app.route("/api/analyze", methods=["POST"])
def api_analyze():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    try:
        data = analyze_excel(f.stream)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/process", methods=["POST"])
def api_process():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    try:
        output, stats = process_excel_with_pagebreaks(f.stream, f.filename)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"processed_{f.filename}",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    import sys
    print(f"Python: {sys.executable}")
    print("openpyxl: ", end="")
    try:
        import openpyxl
        print(f"v{openpyxl.__version__}")
    except ImportError:
        print("NOT installed - run: pip install openpyxl")
    print()
    print("Starting server at http://127.0.0.1:5188")
    print("API endpoints: /api/health, /api/analyze, /api/process")
    app.run(host="127.0.0.1", port=5188, debug=True)
