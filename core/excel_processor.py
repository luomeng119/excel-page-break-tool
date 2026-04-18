"""Excel处理核心 - V1 版本"""
from typing import List, Callable, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.pagebreak import Break
from datetime import datetime
import os
import shutil

from config import OUTPUT_DIR, OUTPUT_TEMPLATE


class ExcelProcessor:
    """Excel分页符处理器"""

    def __init__(self, mark_color: str = 'FF0000'):
        self.progress_callback: Optional[Callable[[int, int], None]] = None
        self.log_callback: Optional[Callable[[str], None]] = None
        self.mark_color = mark_color
        self.preview_mode = False  # 预览模式（带颜色），False为输出模式（不带颜色）

    def set_progress_callback(self, callback: Callable[[int, int], None]):
        self.progress_callback = callback

    def set_log_callback(self, callback: Callable[[str], None]):
        self.log_callback = callback

    def set_mark_color(self, color: str):
        self.mark_color = color

    def set_preview_mode(self, preview: bool):
        self.preview_mode = preview

    def _get_fill(self) -> PatternFill:
        return PatternFill(
            start_color=self.mark_color,
            end_color=self.mark_color,
            fill_type='solid'
        )

    def _log(self, message: str):
        if self.log_callback:
            self.log_callback(message)

    def _backup_original(self, input_path: str) -> Optional[str]:
        """备份原始文件"""
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        ext = os.path.splitext(input_path)[1]
        dir_path = os.path.dirname(input_path)
        backup_name = f"{base_name}_原始备份{ext}"
        backup_path = os.path.join(dir_path, backup_name)
        try:
            shutil.copy2(input_path, backup_path)
            self._log(f"  原始文件已备份: {os.path.basename(backup_path)}")
            return backup_path
        except Exception as e:
            self._log(f"  备份失败: {e}")
            return None

    def process_file(self, input_path: str, output_dir: str = None,
                     skip_backup: bool = False) -> str:
        """
        处理单个Excel文件

        Args:
            input_path: 输入文件路径
            output_dir: 输出目录（默认与输入文件同目录）
            skip_backup: 是否跳过备份（预览模式下跳过）
        """
        self._log(f"开始处理: {os.path.basename(input_path)}")

        # 备份原始文件
        if not skip_backup:
            self._backup_original(input_path)

        # 加载工作簿
        wb = load_workbook(input_path)
        ws = wb.active

        # 清除原有分页符（避免原始文件自带的分页符干扰）
        ws.row_breaks = ws.row_breaks.__class__()

        # 检测序号行
        from core.serial_number_detector import SerialNumberDetector
        detector = SerialNumberDetector()
        serial_rows = detector.detect(ws)

        if not serial_rows:
            self._log("  未检测到序号行，跳过")
            wb.close()
            return None

        self._log(f"  检测到 {len(serial_rows)} 个序号行")

        # 用于去重：记录已添加的分页符 row id（openpyxl row_breaks 不自动去重）
        seen_break_ids = set()
        total_rows = len(serial_rows)
        fill = self._get_fill()
        break_count = 0

        for idx, row_info in enumerate(serial_rows):
            curr_row = row_info.row_index
            prev_row = curr_row - 1

            # 获取上一行的序号值
            prev_val = ws.cell(row=prev_row, column=1).value if prev_row >= 1 else None
            try:
                prev_seq = int(str(prev_val).strip()) if prev_val is not None else None
            except (ValueError, TypeError):
                prev_seq = None  # 无法转成数字（表头文字、空格等）

            # 判断是否需要插入分页符：
            # 关键逻辑：只有"上一行是序号=1的行"时，才认为同卷连续，不插分页符
            # 否则（上一行不是序号=1，或者根本不是数字），都需要分页
            # 特例：第1个序号行如果是文件最开始的数据行，即使上一行是表头也不插
            is_first = (idx == 0)
            is_prev_serial = (prev_seq == 1)

            if is_first:
                # 第1个序号行：文件开头，不需要分页符
                pass
            elif not is_prev_serial:
                # 上一行不是序号1 -> 新卷开始，插入分页符
                break_id = prev_row
                if break_id not in seen_break_ids:
                    ws.row_breaks.append(Break(id=break_id))
                    seen_break_ids.add(break_id)
                    break_count += 1

            # 预览模式下标记背景色
            if self.preview_mode:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=curr_row, column=col)
                    cell.fill = fill

            # 更新进度
            if self.progress_callback:
                self.progress_callback(idx + 1, total_rows)

        self._log(f"  实际插入 {break_count} 个分页符")

        # 生成输出路径
        output_path = self._generate_output_path(input_path, output_dir)

        # 保存
        wb.save(output_path)
        wb.close()

        self._log(f"  完成: {os.path.basename(output_path)}")

        return output_path

    def process_files(self, file_list: List[str], output_dir: str = None) -> List[str]:
        """批量处理多个文件"""
        results = []
        total_files = len(file_list)

        for file_idx, file_path in enumerate(file_list):
            self._log(f"[{file_idx + 1}/{total_files}] 处理中...")
            try:
                output_path = self.process_file(file_path, output_dir)
                if output_path:
                    results.append(output_path)
            except Exception as e:
                self._log(f"  错误: {str(e)}")

        return results

    def _generate_output_path(self, input_path: str, output_dir: str = None) -> str:
        """生成输出文件路径"""
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        date_str = datetime.now().strftime('%Y%m%d')

        from config import APP_VERSION
        # APP_VERSION 可能是 "V1.0"，去掉前导 V 得到纯版本号用于文件名拼接
        version_str = APP_VERSION.lstrip('V')
        file_name = OUTPUT_TEMPLATE.format(filename=base_name, date=date_str, version=version_str)

        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
            return os.path.join(output_dir, file_name)
        else:
            dir_path = os.path.dirname(input_path)
            return os.path.join(dir_path, file_name)
