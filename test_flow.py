"""流程测试脚本 - 不依赖GUI，直接测试核心流程"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from core.serial_number_detector import SerialNumberDetector
from core.excel_processor import ExcelProcessor
from openpyxl import load_workbook
from config import OUTPUT_DIR


def test_file_detection():
    """测试文件检测"""
    test_file = "test_data/卷内目录_原始备份.xlsx"

    print("=" * 60)
    print("【测试1】文件存在性检测")
    print("=" * 60)

    if not os.path.exists(test_file):
        print(f"❌ 测试文件不存在: {test_file}")
        return False

    print(f"✅ 测试文件存在: {test_file}")
    print(f"   文件大小: {os.path.getsize(test_file)} bytes")
    return True


def test_workbook_load():
    """测试工作簿加载"""
    test_file = "test_data/卷内目录_原始备份.xlsx"

    print("\n" + "=" * 60)
    print("【测试2】工作簿加载")
    print("=" * 60)

    try:
        wb = load_workbook(test_file, data_only=True)
        ws = wb.active
        print(f"✅ 工作簿加载成功")
        print(f"   工作表名称: {ws.title}")
        print(f"   总行数: {ws.max_row}")
        print(f"   总列数: {ws.max_column}")
        wb.close()
        return True
    except Exception as e:
        print(f"❌ 工作簿加载失败: {e}")
        return False


def test_serial_detection():
    """测试序号检测"""
    test_file = "test_data/卷内目录_原始备份.xlsx"

    print("\n" + "=" * 60)
    print("【测试3】序号行检测")
    print("=" * 60)

    try:
        wb = load_workbook(test_file, data_only=True)
        ws = wb.active

        detector = SerialNumberDetector()
        results = detector.detect(ws)

        print(f"✅ 序号检测完成")
        print(f"   检测到 {len(results)} 个序号行")

        for i, row_info in enumerate(results[:5]):
            print(f"   行{row_info.row_index}: 值={row_info.value}, 格式={row_info.format_code}, 位数={row_info.display_digits}")

        if len(results) > 5:
            print(f"   ... 还有 {len(results) - 5} 个")

        wb.close()
        return len(results) > 0
    except Exception as e:
        print(f"❌ 序号检测失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_page_break_logic():
    """测试分页符逻辑：验证只有 prev_seq != 1 时才插入分页符"""
    test_file = "test_data/卷内目录_原始备份.xlsx"

    print("\n" + "=" * 60)
    print("【测试4】分页符逻辑验证")
    print("=" * 60)

    try:
        wb = load_workbook(test_file, data_only=True)
        ws = wb.active

        detector = SerialNumberDetector()
        serial_rows = detector.detect(ws)

        # 手动计算应该有多少个分页符（与 excel_processor.py 逻辑一致）
        expected_breaks = 0
        for idx, row_info in enumerate(serial_rows):
            prev_row_idx = row_info.row_index - 1
            if prev_row_idx >= 1:
                prev_cell = ws.cell(row=prev_row_idx, column=1)
                prev_val = prev_cell.value
                try:
                    prev_seq = int(str(prev_val).strip())
                except (ValueError, TypeError):
                    prev_seq = None
                # 与 excel_processor.py 逻辑一致：
                # - 第1个序号行（idx=0）不插分页符
                # - 后续行仅在 prev_seq != 1 时插分页符
                if idx == 0:
                    pass  # 第一个序号行不插
                elif prev_seq != 1:
                    expected_breaks += 1

        print(f"   序号行总数: {len(serial_rows)}")
        print(f"   预期分页符数量: {expected_breaks}")
        print(f"   （第1卷序号1不插分页符，后续卷序号1才插）")

        # 第一个序号行信息
        if serial_rows:
            first = serial_rows[0]
            print(f"   第1个序号行: row={first.row_index}, value={first.value}, format={first.format_code}")
            prev_cell = ws.cell(row=first.row_index - 1, column=1)
            print(f"   该行上一格: value={prev_cell.value} -> 不应为分页点")

        wb.close()

        # 实际处理一个副本，验证插入数量
        import shutil
        test_copy = "/tmp/test_page_break_logic.xlsx"
        shutil.copy2(test_file, test_copy)

        processor = ExcelProcessor(mark_color='FF0000')
        processor.set_preview_mode(False)

        output_path = processor.process_file(test_copy, skip_backup=True)

        wb2 = load_workbook(output_path)
        ws2 = wb2.active
        actual_breaks = len(ws2.row_breaks.brk)
        print(f"   实际写入分页符数量: {actual_breaks}")
        wb2.close()

        if actual_breaks == expected_breaks:
            print(f"✅ 分页符数量正确: {actual_breaks}")
            return True
        else:
            print(f"❌ 分页符数量不符: 预期 {expected_breaks}, 实际 {actual_breaks}")
            return False

    except Exception as e:
        print(f"❌ 分页符逻辑测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_backup():
    """测试原始文件备份功能"""
    import shutil

    test_file = "test_data/卷内目录_原始备份.xlsx"
    test_copy = "/tmp/test_backup_source.xlsx"
    shutil.copy2(test_file, test_copy)

    print("\n" + "=" * 60)
    print("【测试5】原始文件备份功能")
    print("=" * 60)

    try:
        processor = ExcelProcessor()
        processor.set_preview_mode(False)

        backup_path = processor._backup_original(test_copy)
        print(f"   预期备份路径: {backup_path}")

        result = processor.process_file(test_copy, skip_backup=False)

        if os.path.exists(backup_path):
            print(f"✅ 备份文件生成成功: {os.path.basename(backup_path)}")
            print(f"   备份文件大小: {os.path.getsize(backup_path)} bytes")
            return True
        else:
            print(f"❌ 备份文件未生成")
            return False

    except Exception as e:
        print(f"❌ 备份测试失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_output_naming():
    """测试输出文件命名规范"""
    print("\n" + "=" * 60)
    print("【测试6】输出文件命名规范")
    print("=" * 60)

    from config import OUTPUT_TEMPLATE

    # version 输入值应为 "1.0"（不含V），因为 processor 里会 lstrip('V')
    test_cases = [
        ("卷内目录", "20260417", "1.0", "卷内目录_20260417_V1.0.xlsx"),
        ("test_file", "20250417", "1.0", "test_file_20250417_V1.0.xlsx"),
        ("my document", "20250417", "1.0", "my document_20250417_V1.0.xlsx"),
    ]

    print(f"   命名模板: {OUTPUT_TEMPLATE}")

    all_pass = True
    for filename, date, version, expected in test_cases:
        result = OUTPUT_TEMPLATE.format(filename=filename, date=date, version=version)
        status = "✅" if result == expected else "❌"
        print(f"   {status} {filename} -> {result}" + (f" (预期: {expected})" if result != expected else ""))
        if result != expected:
            all_pass = False

    if all_pass:
        print(f"✅ 命名规范检查通过")
    return all_pass


def test_full_process():
    """完整处理流程测试"""
    import shutil

    test_file = "test_data/卷内目录_原始备份.xlsx"
    test_copy = "/tmp/test_full_process.xlsx"
    shutil.copy2(test_file, test_copy)

    print("\n" + "=" * 60)
    print("【测试7】完整处理流程")
    print("=" * 60)

    try:
        processor = ExcelProcessor(mark_color='FF0000')
        processor.set_preview_mode(False)

        progress_log = []

        def log_callback(msg):
            print(f"   [LOG] {msg}")

        def progress_callback(current, total):
            if current % 50 == 0 or current == total:
                print(f"   [进度] {current}/{total}")

        processor.set_log_callback(log_callback)
        processor.set_progress_callback(progress_callback)

        output_path = processor.process_file(test_copy, skip_backup=False)

        if not output_path or not os.path.exists(output_path):
            print(f"❌ 处理失败或未生成输出文件")
            return False

        print(f"✅ 处理完成")
        print(f"   输出文件: {output_path}")
        print(f"   输出大小: {os.path.getsize(output_path)} bytes")

        # 验证输出文件
        wb = load_workbook(output_path)
        ws = wb.active
        print(f"   输出文件行数: {ws.max_row}")
        print(f"   分页符数量: {len(ws.row_breaks.brk)}")

        # 验证备份文件
        backup_path = processor._backup_original(test_copy)
        if os.path.exists(backup_path):
            print(f"   备份文件: ✅ 存在 ({os.path.basename(backup_path)})")
        else:
            print(f"   备份文件: ❌ 不存在")

        wb.close()

        return True

    except Exception as e:
        print(f"❌ 处理流程失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """主测试流程"""
    print("\n" + "=" * 60)
    print("Excel分页符工具 V1 - 完整流程测试")
    print("=" * 60)
    print(f"测试文件: test_data/卷内目录_原始备份.xlsx")
    print(f"输出目录: {OUTPUT_DIR}")
    print("")

    results = []

    results.append(("文件存在性", test_file_detection()))
    results.append(("工作簿加载", test_workbook_load()))
    results.append(("序号检测", test_serial_detection()))
    results.append(("分页符逻辑", test_page_break_logic()))
    results.append(("备份功能", test_backup()))
    results.append(("输出命名", test_output_naming()))
    results.append(("完整流程", test_full_process()))

    # 汇总结果
    print("\n" + "=" * 60)
    print("【测试结果汇总】")
    print("=" * 60)

    passed = 0
    failed = 0

    for name, result in results:
        status = "✅ 通过" if result else "❌ 失败"
        print(f"   {status} - {name}")
        if result:
            passed += 1
        else:
            failed += 1

    print("")
    print(f"总计: {passed}/{len(results)} 通过")

    if failed == 0:
        print("\n🎉 所有测试通过！V1 功能完整实现。")
    else:
        print(f"\n⚠️  有 {failed} 个测试未通过，请检查")

    return failed == 0


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
