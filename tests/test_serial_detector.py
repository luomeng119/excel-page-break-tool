"""序号检测器单元测试"""
import unittest
import sys
import os

# 添加项目目录到路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.serial_number_detector import SerialNumberDetector
from openpyxl import Workbook


class TestSerialNumberDetector(unittest.TestCase):
    """序号检测器测试类"""
    
    def setUp(self):
        """测试前准备"""
        self.detector = SerialNumberDetector()
        self.wb = Workbook()
        self.ws = self.wb.active
    
    def tearDown(self):
        """测试后清理"""
        self.wb.close()
    
    def test_detect_2digit(self):
        """测试2位序号识别 (01)"""
        self.ws['A1'] = 1
        self.ws['A1'].number_format = '00'
        
        results = self.detector.detect(self.ws)
        
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0].row_index, 1)
        self.assertEqual(results[0].display_digits, 2)
    
    def test_detect_3digit(self):
        """测试3位序号识别 (001)"""
        self.ws['A1'] = 1
        self.ws['A1'].number_format = '000'
        
        results = self.detector.detect(self.ws)
        
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0].row_index, 1)
        self.assertEqual(results[0].display_digits, 3)
    
    def test_detect_4digit(self):
        """测试4位序号识别 (0001)"""
        self.ws['A1'] = 1
        self.ws['A1'].number_format = '0000'
        
        results = self.detector.detect(self.ws)
        
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0].row_index, 1)
        self.assertEqual(results[0].display_digits, 4)
    
    def test_ignore_normal_number(self):
        """测试普通数字不识别"""
        self.ws['A1'] = 1
        self.ws['A1'].number_format = 'General'
        
        results = self.detector.detect(self.ws)
        
        self.assertEqual(len(results), 0)
    
    def test_detect_multiple_rows(self):
        """测试多行序号检测"""
        # 第1行 - 序号行
        self.ws['A1'] = 1
        self.ws['A1'].number_format = '000'
        
        # 第2行 - 普通数据
        self.ws['A2'] = '数据'
        
        # 第3行 - 序号行
        self.ws['A3'] = 1
        self.ws['A3'].number_format = '000'
        
        results = self.detector.detect(self.ws)
        
        self.assertEqual(len(results), 2)
        self.assertEqual(results[0].row_index, 1)
        self.assertEqual(results[1].row_index, 3)
    
    def test_ignore_non_first_column(self):
        """测试非第一列不识别"""
        self.ws['B1'] = 1
        self.ws['B1'].number_format = '000'
        
        results = self.detector.detect(self.ws)
        
        self.assertEqual(len(results), 0)


if __name__ == '__main__':
    unittest.main()
