"""右侧预览面板"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QScrollArea,
    QFrame, QSizePolicy
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QPainter, QColor, QPen, QFont, QBrush
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os


class PreviewPanel(QWidget):
    """预览面板"""

    def __init__(self):
        super().__init__()
        self.current_page = 1
        self.total_pages = 1
        self.zoom_level = 100
        self.excel_path = None
        self.page_breaks = []
        self.serial_rows = []
        self.highlight_color = 'FF0000'
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # 工具栏
        toolbar = QHBoxLayout()

        self.zoom_out_btn = QPushButton("-")
        self.zoom_out_btn.setFixedWidth(40)
        self.zoom_out_btn.setToolTip("缩小")
        self.zoom_out_btn.clicked.connect(self._zoom_out)

        self.zoom_label = QLabel("100%")
        self.zoom_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.zoom_label.setFixedWidth(60)
        self.zoom_label.setStyleSheet("font-weight: bold;")

        self.zoom_in_btn = QPushButton("+")
        self.zoom_in_btn.setFixedWidth(40)
        self.zoom_in_btn.setToolTip("放大")
        self.zoom_in_btn.clicked.connect(self._zoom_in)

        toolbar.addWidget(self.zoom_out_btn)
        toolbar.addWidget(self.zoom_label)
        toolbar.addWidget(self.zoom_in_btn)
        toolbar.addStretch()

        self.prev_btn = QPushButton("上一页")
        self.prev_btn.setEnabled(False)
        self.prev_btn.clicked.connect(self._prev_page)

        self.page_label = QLabel("第 1/1 页")
        self.page_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.page_label.setFixedWidth(100)
        self.page_label.setStyleSheet("font-weight: bold;")

        self.next_btn = QPushButton("下一页")
        self.next_btn.setEnabled(False)
        self.next_btn.clicked.connect(self._next_page)

        toolbar.addWidget(self.prev_btn)
        toolbar.addWidget(self.page_label)
        toolbar.addWidget(self.next_btn)

        layout.addLayout(toolbar)

        # 预览画布
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("background-color: #f0f0f0;")

        self.canvas = PreviewCanvas()
        scroll.setWidget(self.canvas)

        layout.addWidget(scroll)

        # 信息标签
        self.info_label = QLabel("请选择文件进行预览")
        self.info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.info_label)

    def load_file(self, file_path):
        """加载原始文件进行预览"""
        if not file_path or not os.path.exists(file_path):
            return

        self.excel_path = file_path
        self.info_label.setText(f"原始文件: {os.path.basename(file_path)}")

        try:
            self.canvas.load_excel(file_path, show_processed=False)
            self.page_breaks = self.canvas.page_breaks
            self.serial_rows = self.canvas.serial_rows

            self.total_pages = max(1, len(self.page_breaks) + 1)
            self.current_page = 1
            self._update_page_controls()

        except Exception as e:
            self.info_label.setText(f"加载失败: {str(e)}")

    def load_processed_file(self, file_path):
        """加载处理后的文件进行预览"""
        if not file_path or not os.path.exists(file_path):
            return

        self.excel_path = file_path
        self.info_label.setText(f"处理后: {os.path.basename(file_path)}")

        try:
            self.canvas.load_excel(file_path, show_processed=True)
            self.page_breaks = self.canvas.page_breaks
            self.serial_rows = self.canvas.serial_rows

            self.total_pages = max(1, len(self.page_breaks) + 1)
            self.current_page = 1
            self._update_page_controls()

        except Exception as e:
            self.info_label.setText(f"加载失败: {str(e)}")

    def _zoom_in(self):
        self.zoom_level = min(200, self.zoom_level + 20)
        self.zoom_label.setText(f"{self.zoom_level}%")
        self.canvas.set_zoom(self.zoom_level)

    def _zoom_out(self):
        self.zoom_level = max(50, self.zoom_level - 20)
        self.zoom_label.setText(f"{self.zoom_level}%")
        self.canvas.set_zoom(self.zoom_level)

    def _prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self._update_page_controls()
            self.canvas.set_page(self.current_page)

    def _next_page(self):
        if self.current_page < self.total_pages:
            self.current_page += 1
            self._update_page_controls()
            self.canvas.set_page(self.current_page)

    def _update_page_controls(self):
        self.page_label.setText(f"第 {self.current_page}/{self.total_pages} 页")
        self.prev_btn.setEnabled(self.current_page > 1)
        self.next_btn.setEnabled(self.current_page < self.total_pages)

    def set_highlight_color(self, color_hex: str):
        """设置高亮颜色（仅重绘，不重新加载文件）"""
        self.highlight_color = color_hex
        self.canvas.set_highlight_color(color_hex)


class PreviewCanvas(QWidget):
    """预览画布 - 自定义绘制"""

    def __init__(self):
        super().__init__()
        self.zoom = 1.0
        self.current_page = 1
        self.excel_data = []
        self.headers = []
        self.page_breaks = []
        self.serial_rows = []
        self.highlight_color = 'FF0000'
        self.setMinimumSize(600, 800)
        self.setSizePolicy(
            QSizePolicy.Policy.Expanding,
            QSizePolicy.Policy.Expanding
        )

    def set_highlight_color(self, color_hex: str):
        """设置高亮颜色并立即重绘"""
        self.highlight_color = color_hex
        self.update()

    def _hex_to_qcolor(self, hex_color: str) -> QColor:
        """将 hex 颜色转为 QColor"""
        hex_color = hex_color.lstrip('#')
        if len(hex_color) >= 6:
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)
            return QColor(r, g, b)
        return QColor(255, 200, 200)

    def load_excel(self, file_path, show_processed=False):
        """加载Excel数据"""
        self.excel_data = []
        self.headers = []
        self.page_breaks = []
        self.serial_rows = []

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # 读取表头（前两行）
        for row_idx in range(1, min(3, ws.max_row + 1)):
            for col in range(1, min(ws.max_column + 1, 7)):
                cell = ws.cell(row=row_idx, column=col)
                if row_idx == 2:  # 第二行作为列名
                    self.headers.append(str(cell.value) if cell.value else f"列{col}")

        # 如果没有读到表头，用第一行
        if not self.headers:
            for col in range(1, min(ws.max_column + 1, 7)):
                cell = ws.cell(row=1, column=col)
                self.headers.append(str(cell.value) if cell.value else f"列{col}")

        # 检测序号行
        from core.serial_number_detector import SerialNumberDetector
        detector = SerialNumberDetector()
        detected_rows = detector.detect(ws)
        self.serial_rows = [r.row_index for r in detected_rows]

        # 读取数据（上限500行）
        max_rows = min(ws.max_row + 1, 502)

        for row_idx in range(3, max_rows):
            row_data = []

            is_colored_background = False
            bg_color = None
            if show_processed:
                first_cell = ws.cell(row=row_idx, column=1)
                if first_cell.fill and first_cell.fill.start_color:
                    color = first_cell.fill.start_color.rgb
                    if color and str(color) not in ['00000000', 'FFFFFFFF', 'None', '']:
                        is_colored_background = True
                        bg_color = str(color)

            for col in range(1, min(ws.max_column + 1, 7)):
                cell = ws.cell(row=row_idx, column=col)
                value = cell.value
                if value is None:
                    value = ""
                elif isinstance(value, (int, float)):
                    if col == 1 and row_idx in self.serial_rows:
                        num_format = cell.number_format or ''
                        if '0000' in num_format:
                            value = f"{int(value):04d}"
                        elif '000' in num_format:
                            value = f"{int(value):03d}"
                        elif '00' in num_format:
                            value = f"{int(value):02d}"
                row_data.append(str(value))

            is_serial = row_idx in self.serial_rows

            current_page_rows_count = len(self.excel_data)
            self.excel_data.append({
                'row_index': row_idx,
                'data': row_data,
                'is_serial': is_serial,
                'is_colored_background': is_colored_background,
                'bg_color': bg_color
            })

            # 分页点：序号行且不是第一个
            if is_serial and current_page_rows_count > 0:
                self.page_breaks.append(current_page_rows_count - 1)

        wb.close()
        self.update()

    def set_zoom(self, percentage):
        self.zoom = percentage / 100.0
        self.update()

    def set_page(self, page_num):
        self.current_page = page_num
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # 背景
        painter.fillRect(self.rect(), QColor(240, 240, 240))

        # 纸张
        paper_width = int(500 * self.zoom)
        paper_height = int(700 * self.zoom)
        paper_x = max(20, (self.width() - paper_width) // 2)
        paper_y = 20

        # 阴影
        painter.fillRect(paper_x + 3, paper_y + 3, paper_width, paper_height, QColor(200, 200, 200))
        # 白纸
        painter.fillRect(paper_x, paper_y, paper_width, paper_height, Qt.GlobalColor.white)
        painter.setPen(QPen(QColor(180, 180, 180), 1))
        painter.drawRect(paper_x, paper_y, paper_width, paper_height)

        if not self.excel_data:
            painter.setPen(QColor(150, 150, 150))
            painter.setFont(QFont("Arial", int(12 * self.zoom)))
            painter.drawText(paper_x, paper_y, paper_width, paper_height,
                             Qt.AlignmentFlag.AlignCenter, "暂无预览数据")
            painter.end()
            return

        margin = int(20 * self.zoom)
        content_x = paper_x + margin
        content_y = paper_y + margin
        content_width = paper_width - 2 * margin

        row_height = int(25 * self.zoom)
        col_width = content_width // max(len(self.headers), 1)

        # 表头
        header_font = QFont("Arial", int(10 * self.zoom))
        header_font.setBold(True)
        painter.setFont(header_font)
        painter.setPen(QPen(QColor(100, 100, 100), 1))

        for col_idx, header in enumerate(self.headers):
            x = content_x + col_idx * col_width
            painter.drawText(x, content_y, col_width - 5, row_height,
                             Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                             header[:10])

        content_y += row_height
        painter.setPen(QPen(QColor(200, 200, 200), 1))
        painter.drawLine(content_x, content_y, content_x + content_width, content_y)

        # 数据行
        data_font = QFont("Arial", int(9 * self.zoom))
        painter.setFont(data_font)

        # 使用用户选择的颜色（highlight_color）
        highlight_qcolor = self._hex_to_qcolor(self.highlight_color)
        # 生成浅色背景（降低饱和度用于背景填充）
        light_qcolor = QColor(
            min(255, highlight_qcolor.red() // 3 + 200),
            min(255, highlight_qcolor.green() // 3 + 180),
            min(255, highlight_qcolor.blue() // 3 + 180),
        )

        for row_idx, row_data in enumerate(self.excel_data):
            content_y += row_height

            if content_y > paper_y + paper_height - margin:
                break

            # 分页线
            if row_idx in self.page_breaks:
                line_y = content_y - row_height + 5
                if line_y > paper_y + margin:
                    painter.setPen(QPen(QColor(255, 0, 0), 3, Qt.PenStyle.DashLine))
                    painter.drawLine(content_x - 10, line_y, content_x + content_width + 10, line_y)
                    painter.setPen(QColor(255, 0, 0))
                    painter.setFont(QFont("Arial", int(8 * self.zoom)))
                    painter.drawText(content_x + content_width - 50, line_y - 15, 50, 15,
                                     Qt.AlignmentFlag.AlignRight, "分页")

            # 序号行高亮（使用用户选择的颜色）
            is_serial = row_data.get('is_serial', False)
            is_colored = row_data.get('is_colored_background', False)

            if is_serial or is_colored:
                # 优先使用文件中的实际背景色，否则使用用户选择的高亮色
                if is_colored and row_data.get('bg_color'):
                    bg_hex = row_data['bg_color']
                    if len(bg_hex) >= 6:
                        r = int(bg_hex[0:2], 16)
                        g = int(bg_hex[2:4], 16)
                        b = int(bg_hex[4:6], 16)
                        fill_color = QColor(r, g, b)
                    else:
                        fill_color = light_qcolor
                else:
                    # 使用用户选择的颜色的浅色版本
                    fill_color = light_qcolor

                painter.fillRect(content_x - 5, content_y - row_height + 2,
                                 content_width + 10, row_height - 2, fill_color)
                painter.setPen(QPen(highlight_qcolor.darker(150), 1))
            else:
                painter.setPen(QPen(QColor(80, 80, 80), 1))

            # 单元格内容
            for col_idx, cell_value in enumerate(row_data['data'][:len(self.headers)]):
                x = content_x + col_idx * col_width
                display_text = str(cell_value)[:12] if cell_value else ""
                painter.drawText(x, content_y - row_height + 5, col_width - 5, row_height - 5,
                                 Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                                 display_text)

        painter.end()
